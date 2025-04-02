"""
Module: get_size_data

This module provides functions to handle data extraction and processing from Excel files,
as well as downloading and extracting files from the internet.
"""

import json
import logging
import os
import sys
import zipfile

import openpyxl
import requests


def read_xlsx_to_dict(
    file_path: str, sheet_name: str, columns: list = None
) -> dict:
    """
    Read data from an Excel file and convert it to a dictionary.

    Args:
        file_path (str): The path to the Excel file.
        sheet_name (str): The name of the sheet to read.
        columns (list, optional): A list of columns to include in the dictionary. Defaults to None.

    Returns:
        dict: The data read from the Excel file as a dictionary.
    """
    try:
        # Load the workbook and select the specified sheet
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
    except FileNotFoundError:
        logging.error("The file '%s' was not found.", file_path)
        sys.exit(1)
    except KeyError:
        logging.error(
            "Error: The sheet '%s' does not exist in the workbook.", sheet_name
        )
        sys.exit(1)

    # Initialize an empty dictionary to store the data
    data_dict = {}

    # Read the header row to map column names to indices
    header = list(sheet.iter_rows(values_only=True))[0]
    column_indices = {name: index for index, name in enumerate(header)}

    # Filter the columns based on the provided list of column names
    if columns:
        selected_indices = [
            column_indices[name] for name in columns if name in column_indices
        ]
    else:
        selected_indices = list(
            range(1, len(header))
        )  # Default to all columns except the first one

    if len(selected_indices) != len(columns):
        logging.error(
            "Some columns were not found in the sheet. "
            "Columns found: %s, Columns expected: %s",
            [header[i] for i in selected_indices],
            columns,
        )
        sys.exit(1)

    if len(selected_indices) > 1:
        # Iterate through the rows and columns to read data

        for row in list(sheet.iter_rows(values_only=True))[1:]:
            values = tuple(
                row[i] for i in selected_indices
            )  # Filter columns based on the provided list
            single_species = {}
            i = 1
            species_key = values[0]
            for index in selected_indices[1:]:
                single_species[header[index]] = values[i]
                i += 1
            data_dict[species_key] = single_species
    else:
        logging.error("There must be at least 2 columns to read.")
        sys.exit(1)
    return data_dict


def download_file(url: str, dest_path: str):
    """
    Download a file from a given URL and save it to the specified destination path.

    Args:
        url (str): The URL of the file to download.
        dest_path (str): The destination path to save the downloaded file.
    """
    # Function implementation goes here
    try:
        # Make HTTP request to download the file
        response = requests.get(url, timeout=30)
        # Raise an exception if the request was unsuccessful
        response.raise_for_status()
    except requests.exceptions.RequestException as e:
        logging.error(
            "Failed to download the file from %s. Reason: %s", url, str(e)
        )
        sys.exit(1)

    try:
        # Write the content to the specified file path
        with open(dest_path, "wb") as file:
            file.write(response.content)
    except IOError as e:
        logging.error(
            "Error: Failed to write the file to %s. Reason: %s",
            dest_path,
            str(e),
        )
        sys.exit(1)


def extract_file_from_zip(zip_path: str, file_name: str, dest_dir: str):
    """
    Extract a file from a ZIP archive to the specified destination directory.

    Args:
        zip_path (str): The path to the ZIP archive.
        file_name (str): The name of the file to extract.
        dest_dir (str): The destination directory to extract the file to.
    """
    # Function implementation goes here
    try:
        # Ensure the destination directory exists
        os.makedirs(dest_dir, exist_ok=True)

        # Open the ZIP file
        with zipfile.ZipFile(zip_path, "r") as zip_ref:
            # Extract the specified file
            zip_ref.extract(file_name, dest_dir)
            logging.info("File '%s' extracted to '%s'", file_name, dest_dir)
    except FileNotFoundError:
        logging.error("The file '%s' was not found.", zip_path)
        sys.exit(1)

    except KeyError:
        logging.error(
            "Error: The file '%s' does not exist in the ZIP archive.",
            file_name,
        )
        sys.exit(1)

    except zipfile.BadZipFile:
        logging.error("The file '%s' is not a valid ZIP file.", zip_path)
        sys.exit(1)

    # The following is meant to capture unexpected errors
    except Exception as e:
        logging.error("An unexpected error occurred. Reason: %s", str(e))
        sys.exit(1)


def write_dict_to_json(data: dict, file_path: str):
    """
    Write a dictionary to a JSON file.

    Args:
        data (dict): The dictionary to write.
        file_path (str): The path to the JSON file.

    Returns:
        None
    """
    try:
        # Open the file in write mode
        with open(file_path, mode="wt", encoding="utf-8") as json_file:
            # Write the dictionary to the JSON file
            json.dump(data, json_file, indent=4)
            logging.info("Dictionary successfully written to '%s'", file_path)
    except IOError as e:
        logging.error(
            "Error: Failed to write to the file '%s'. Reason: %s",
            file_path,
            str(e),
        )


def read_dict_from_json(file_path: str) -> dict:
    """
    Read a dictionary from a JSON file.

    Args:
        file_path (str): The path to the JSON file.

    Returns:
        dict: The dictionary read from the JSON file.
    """
    try:
        # Open the file in read mode
        with open(file_path, mode="rt", encoding="utf-8") as json_file:
            # Read the JSON file and convert it to a dictionary
            data = json.load(json_file)
            logging.info("Dictionary successfully read from '%s'", file_path)
            return data
    except IOError as e:
        logging.error(
            "Error: Failed to read the file '%s'. Reason: %s",
            file_path,
            str(e),
        )
    except json.JSONDecodeError as e:
        logging.error(
            "Error: Failed to decode JSON from the file '%s'. Reason: %s",
            file_path,
            str(e),
        )
    return {}


def get_new_avonet_data():
    """
    Download new Avonet data from a specified URL.
    """
    url = "https://figshare.com/ndownloader/files/34480856"
    xlsx_path = "temp/aviform_data.xlsx"
    download_file(url, xlsx_path)


def process_avonet_data():
    """
    Process Avonet data to put it in json format.
    """
    xlsx_path = "temp/aviform_data.xlsx"
    sheet_name = "AVONET2_eBird"
    columns = ["Species2", "Wing.Length", "Habitat", "Mass"]
    data = read_xlsx_to_dict(xlsx_path, sheet_name, columns)

    # Write the dictionary to a JSON file
    json_file_path = "temp/aviform_data.json"
    write_dict_to_json(data, json_file_path)


def read_cached_avonet_data() -> dict:
    """
    Read cached Avonet data from a JSON file.

    Returns:
        dict: The cached Avonet data as a dictionary.
    """
    json_file_path = "temp/aviform_data.json"
    return read_dict_from_json(json_file_path)
